#!/usr/bin/env python3
"""
SMR Weekly Stats Automation — fills a .xlsm billing workbook from a text input file.

Usage:
    python3 smr_filler.py <workbook.xlsm> <input.txt>

Produces <workbook_filled.xlsm> with treatment codes, comments, notes, and
daily totals written into the "STANDARD Stats" sheet.
"""

import re
import sys
import os
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SHEET_NAME = "STANDARD Stats"
STUDENT_ROW_START = 6
STUDENT_ROW_END = 49       # inclusive
DATE_COL_START = 7         # G
DATE_COL_END = 22          # V (inclusive — but may be None)
COMMENT_COL = 23           # W
TOTAL_ROW = 101
NOTES_ROW = 131
KNOWN_CODES = {"T", "G1", "G2", "G3", "I", "M", "A", "O", "H", "Y", "S",
               "R", "D", "C"}
NON_BILLABLE = {"A", "O", "H"}
SCHEDULING_CODE = "S"
SCHEDULING_MINUTES = 15
GROUP_CODES = {"G1", "G2", "G3"}

# ---------------------------------------------------------------------------
# Input Parser
# ---------------------------------------------------------------------------

def _parse_grid_sessions(body_lines):
    """Parse grid-format sessions (tab-separated table from Apple Notes)."""
    sessions = []
    dates = []

    for line in body_lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            continue
        # Skip decorative separators (em-dash lines from Apple Notes)
        if all(c in '—-–_' for c in stripped):
            continue

        parts = line.split("\t")

        if not dates:
            # Header row: "name", date1, date2, ...
            dates = [p.strip() for p in parts[1:]]
            continue

        # Data row: student name + codes per date
        student = parts[0].strip()
        if not student:
            continue

        for i, date in enumerate(dates):
            if not date:
                continue
            if i + 1 < len(parts):
                code = parts[i + 1].strip()
                if code:
                    sessions.append((date, student, code))

    return sessions


def parse_input(path):
    """Parse the three-section input file. Returns (sessions, comments, notes)."""
    with open(path) as f:
        text = f.read()

    sections = re.split(r'^##\s*', text, flags=re.MULTILINE)
    sessions, comments, notes = [], [], []

    for section in sections:
        if not section.strip():
            continue
        lines = section.strip().splitlines()
        header = lines[0].strip().lower()
        body_lines = lines[1:]

        if header.startswith("session"):
            # Detect grid format: first non-comment data line contains tabs
            data_lines = [l for l in body_lines
                          if l.strip() and not l.strip().startswith("#")
                          and not all(c in '—-–_' for c in l.strip())]
            if data_lines and "\t" in data_lines[0]:
                sessions = _parse_grid_sessions(body_lines)
            else:
                for line in body_lines:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = [p.strip() for p in line.split(",", 2)]
                    if len(parts) == 3:
                        sessions.append((parts[0], parts[1], parts[2]))
                    else:
                        print(f"  WARN: bad session line: {line!r}")

        elif header.startswith("comment"):
            for line in body_lines:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if all(c in '—-–_' for c in line):
                    continue
                parts = [p.strip() for p in line.split(",", 1)]
                if len(parts) == 2:
                    comments.append((parts[0], parts[1]))
                else:
                    print(f"  WARN: bad comment line: {line!r}")

        elif header.startswith("note"):
            for line in body_lines:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = [p.strip() for p in line.split(",", 1)]
                if len(parts) == 2:
                    notes.append((parts[0], parts[1]))
                else:
                    print(f"  WARN: bad note line: {line!r}")

    return sessions, comments, notes

# ---------------------------------------------------------------------------
# Student Map
# ---------------------------------------------------------------------------

def build_student_map(ws):
    """Build lookup dict from student keys to row numbers.

    Returns (student_map, ambiguous_lastnames).
    student_map maps normalized keys -> row.
    ambiguous_lastnames is a set of last names that appear more than once.
    """
    student_map = {}          # normalized key -> row
    lastname_counts = defaultdict(list)  # lastname -> [row, ...]
    # Track "last first" keys (without suffix) to detect collisions
    # e.g. "Bade, Elijah (a)" and "Bade, Elijah (b)" both produce "bade elijah"
    base_name_counts = defaultdict(list)  # "last first" -> [row, ...]

    for row in range(STUDENT_ROW_START, STUDENT_ROW_END + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        name = str(name).strip()
        name_lower = name.lower()

        # Full cell value key (e.g. "folie, caleb (a)")
        student_map[name_lower] = row

        # Parse "Last, First" or "Last, First (A)"
        m = re.match(r'^([^,]+),\s*(.+)$', name)
        if m:
            last = m.group(1).strip().lower()
            first_full = m.group(2).strip().lower()  # e.g. "caleb (a)"
            first = re.sub(r'\s*\([^)]*\)\s*$', '', first_full).strip()

            base_key = f"{last} {first}"
            base_name_counts[base_key].append(row)

            # "last first (x)" key if suffix present
            if first_full != first:
                student_map[f"{last} {first_full}"] = row

            lastname_counts[last].append(row)

    ambiguous = {ln for ln, rows in lastname_counts.items() if len(rows) > 1}

    # Add "last first" keys only when unambiguous (no suffix collisions)
    for base_key, rows in base_name_counts.items():
        if len(rows) == 1:
            student_map[base_key] = rows[0]

    # Add last-name-only keys for unambiguous names
    for ln, rows in lastname_counts.items():
        if len(rows) == 1:
            student_map[ln] = rows[0]

    return student_map, ambiguous


def resolve_student(key, student_map, ambiguous):
    """Resolve an input student key to a row number, or None."""
    norm = key.strip().lower()
    # Remove commas for flexible matching: "Folie, Caleb (A)" -> "folie caleb (a)"
    norm_no_comma = norm.replace(",", "").strip()
    # Collapse multiple spaces
    norm_no_comma = re.sub(r'\s+', ' ', norm_no_comma)

    # 1. Exact match (full cell value with comma)
    if norm in student_map:
        return student_map[norm]

    # 2. Match without comma
    if norm_no_comma in student_map:
        return student_map[norm_no_comma]

    # 3. Check if it's a last-name only lookup
    parts = norm_no_comma.split()
    if len(parts) == 1:
        if norm_no_comma in ambiguous:
            print(f"  WARN: ambiguous last name '{key}' — specify first name")
            return None
        if norm_no_comma in student_map:
            return student_map[norm_no_comma]

    return None

# ---------------------------------------------------------------------------
# Date Map
# ---------------------------------------------------------------------------

def build_date_map(ws):
    """Build {(month, day): column} from row 5 dates."""
    date_map = {}
    for col in range(DATE_COL_START, DATE_COL_END + 1):
        val = ws.cell(row=5, column=col).value
        if val is None:
            continue
        # val is a datetime object
        if hasattr(val, 'month'):
            date_map[(val.month, val.day)] = col
    return date_map


def resolve_date(date_str, date_map):
    """Parse 'm/d' string and look up column."""
    date_str = date_str.strip().rstrip(":")
    parts = date_str.split("/")
    if len(parts) == 2:
        try:
            m, d = int(parts[0]), int(parts[1])
            return date_map.get((m, d))
        except ValueError:
            pass
    return None

# ---------------------------------------------------------------------------
# Code Cell Parser
# ---------------------------------------------------------------------------

def _extract_codes_greedy(token):
    """Try to extract concatenated codes from a token like 'TT/', 'G1C/', 'DD'.

    Scans left-to-right, matching the longest known code at each position,
    then consuming any trailing '/' or '//' modifier.

    Returns list of (code, modifier) tuples, or None if parsing fails.
    """
    text = token.upper()
    results = []
    i = 0
    # Sort known codes longest-first so G1 is tried before G
    sorted_codes = sorted(KNOWN_CODES, key=len, reverse=True)

    while i < len(text):
        matched = False
        for code in sorted_codes:
            if text[i:i+len(code)] == code:
                j = i + len(code)
                if j < len(text) and text[j] == '/':
                    if j + 1 < len(text) and text[j+1] == '/':
                        results.append((code, 0.25))
                        i = j + 2
                    else:
                        results.append((code, 0.5))
                        i = j + 1
                else:
                    results.append((code, 1.0))
                    i = j
                matched = True
                break
        if not matched:
            return None  # Can't fully parse this token

    return results if results else None


def parse_code_cell(cell_value):
    """Parse a treatment code cell into [(code, modifier), ...].

    Modifier: 1.0 = full, 0.5 = half (/), 0.25 = quarter (//), 2.0 = doubled.
    """
    if not cell_value:
        return []

    text = str(cell_value).strip()
    if not text:
        return []

    results = []

    # Tokenize: split on whitespace, but keep slashes attached
    tokens = text.split()

    i = 0
    while i < len(tokens):
        token = tokens[i]

        # Handle token that is just "/" — modifier for the previous code
        if token == "/":
            # This is a standalone slash — means previous code is half
            # and next code (if any) is also half
            # e.g. "C /T" tokenizes as ["C", "/T"] but
            #      "C / T" tokenizes as ["C", "/", "T"]
            if results:
                # Change last result to half
                code, mod = results[-1]
                results[-1] = (code, 0.5)
            i += 1
            continue

        if token == "//":
            if results:
                code, mod = results[-1]
                results[-1] = (code, 0.25)
            i += 1
            continue

        # Check for trailing slashes: "T/" or "T//"
        if token.endswith("//") and token[:-2].upper() in KNOWN_CODES:
            results.append((token[:-2].upper(), 0.25))
            i += 1
            continue

        if token.endswith("/") and token[:-1].upper() in KNOWN_CODES:
            code = token[:-1].upper()
            # Check if next token starts with a code (mixed half: "T/G1")
            # Actually "T/" is just half T. "T/G1" is a single token.
            results.append((code, 0.5))
            i += 1
            continue

        # Check for slash in middle: "T/G1" — mixed half session
        if "/" in token and not token.startswith("/"):
            slash_parts = token.split("/")
            # Filter empties from trailing slashes
            slash_parts = [p for p in slash_parts if p]
            if all(p.upper() in KNOWN_CODES for p in slash_parts):
                for p in slash_parts:
                    results.append((p.upper(), 0.5))
                i += 1
                continue

        # Check for leading slash: "/T" means the code is half
        if token.startswith("/") and token[1:].upper() in KNOWN_CODES:
            results.append((token[1:].upper(), 0.5))
            i += 1
            continue

        # Plain code
        upper = token.upper()
        if upper in KNOWN_CODES:
            results.append((upper, 1.0))
            i += 1
            continue

        # Unknown token — try greedy extraction of concatenated codes
        # e.g. "TT/" → T + T/, "G1C/" → G1 + C/, "DD" → D + D
        greedy = _extract_codes_greedy(token)
        if greedy:
            results.extend(greedy)
            i += 1
            continue

        # Truly unknown — pass through but warn
        print(f"  WARN: unknown code token '{token}' in cell '{text}'")
        results.append((token.upper(), 1.0))
        i += 1

    # Post-process: detect doubled codes (same code appearing twice = 2x session)
    # e.g. "T T" -> [("T", 1.0), ("T", 1.0)] — each is a full session
    # This is the correct behavior already; no modification needed.

    return results

# ---------------------------------------------------------------------------
# Duration Parsing
# ---------------------------------------------------------------------------

def parse_duration_minutes(duration_str):
    """Parse '30 min' -> 30.  Returns int minutes."""
    if not duration_str:
        return 30  # default
    m = re.search(r'(\d+)', str(duration_str))
    return int(m.group(1)) if m else 30


DURATION_RE = re.compile(
    r'\((\d+\.?\d*)\s*(min|mins|minutes|minute|mn|m|hr|hrs|hour|hours|h)\s*\)',
    re.IGNORECASE
)

def parse_notes_durations(notes_text, month, day):
    """Extract total minutes from the notes section for a specific date.

    Finds the date header (e.g. "10/1") and parses all (Xmin)/(Xhr) durations
    from that date's section until the next date header or end of text.
    """
    if not notes_text:
        return 0

    # Build pattern for this date — match "10/1" or "10/1:" at start of line
    date_pattern = re.compile(
        rf'^\s*{month}/{day}\s*:?\s*$', re.MULTILINE
    )

    # Find this date's section
    match = date_pattern.search(notes_text)
    if not match:
        return 0

    # Find the next date header or end
    remaining = notes_text[match.end():]
    next_date = re.search(r'^\s*\d{1,2}/\d{1,2}\s*:?\s*$', remaining, re.MULTILINE)
    if next_date:
        section = remaining[:next_date.start()]
    else:
        section = remaining

    total = 0
    for m in DURATION_RE.finditer(section):
        amount = float(m.group(1))
        unit = m.group(2).lower()
        if unit in ('hr', 'hrs', 'hour', 'hours', 'h'):
            amount *= 60
        total += amount

    return total

# ---------------------------------------------------------------------------
# Fill Functions
# ---------------------------------------------------------------------------

def fill_treatment_codes(ws, sessions, student_map, ambiguous, date_map):
    """Write treatment codes into the grid. Returns count of cells written."""
    count = 0
    for date_str, student_key, code in sessions:
        row = resolve_student(student_key, student_map, ambiguous)
        if row is None:
            print(f"  WARN: student not found: '{student_key}'")
            continue
        col = resolve_date(date_str, date_map)
        if col is None:
            print(f"  WARN: date not found: '{date_str}'")
            continue
        ws.cell(row=row, column=col, value=code)
        count += 1
    return count


def fill_comments(ws, comments, student_map, ambiguous):
    """Write comments into column W. Returns count of students with comments."""
    # Collect all comments per student row so multiple comments concatenate
    row_comments = defaultdict(list)
    for student_key, text in comments:
        row = resolve_student(student_key, student_map, ambiguous)
        if row is None:
            print(f"  WARN: student not found for comment: '{student_key}'")
            continue
        row_comments[row].append(text)

    for row, texts in row_comments.items():
        combined = "; ".join(texts)
        ws.cell(row=row, column=COMMENT_COL, value=combined)

    return len(row_comments)


def fill_notes(ws, notes):
    """Build multi-line notes string and write to A131+. Returns date count."""
    if not notes:
        return 0

    # Group by date, preserving order
    from collections import OrderedDict
    grouped = OrderedDict()
    for date_str, text in notes:
        grouped.setdefault(date_str, []).append(text)

    lines = []
    for date_str, texts in grouped.items():
        lines.append(date_str)
        lines.append(" ".join(texts))

    full_text = "\n".join(lines)

    # Write to A131. If very long, split across A131-A133.
    MAX_CELL = 32000  # Excel cell character limit is 32767
    if len(full_text) <= MAX_CELL:
        ws.cell(row=NOTES_ROW, column=1, value=full_text)
    else:
        chunks = [full_text[i:i+MAX_CELL] for i in range(0, len(full_text), MAX_CELL)]
        for idx, chunk in enumerate(chunks[:3]):
            ws.cell(row=NOTES_ROW + idx, column=1, value=chunk)

    return len(grouped)


def calculate_daily_totals(ws_edit, ws_data, date_map, notes_text):
    """Calculate and write daily totals to row 101.

    ws_edit: the workbook we write to (data_only=False)
    ws_data: the workbook we read computed values from (data_only=True)

    Returns dict of {date_str: total_minutes}.
    """
    totals = {}

    for (month, day), col in date_map.items():
        total_minutes = 0
        seen_groups = set()  # Each group code (G1, G2, G3) counts once per day

        # Step A: Sum session minutes from treatment grid
        for row in range(STUDENT_ROW_START, STUDENT_ROW_END + 1):
            # Read cell from the edit workbook (we may have just written to it)
            cell_val = ws_edit.cell(row=row, column=col).value
            if not cell_val or (isinstance(cell_val, str) and cell_val.startswith("=")):
                # It's a formula or empty — check if data_only has a value
                data_val = ws_data.cell(row=row, column=col).value
                if data_val and not (isinstance(data_val, str) and data_val.startswith("=")):
                    cell_val = data_val
                else:
                    continue

            codes = parse_code_cell(cell_val)
            if not codes:
                continue

            # Get student's base duration from col C
            dur_str = ws_data.cell(row=row, column=3).value
            base_minutes = parse_duration_minutes(dur_str)

            for code, modifier in codes:
                if code.upper() in NON_BILLABLE:
                    continue
                if code.upper() == SCHEDULING_CODE:
                    total_minutes += SCHEDULING_MINUTES
                elif code.upper() in GROUP_CODES:
                    if code.upper() not in seen_groups:
                        seen_groups.add(code.upper())
                        total_minutes += base_minutes * modifier
                else:
                    total_minutes += base_minutes * modifier

        # Step B: Add notes time for this date
        notes_minutes = parse_notes_durations(notes_text, month, day)
        total_minutes += notes_minutes

        # Step C: Write to row 101
        date_str = f"{month}/{day}"
        totals[date_str] = total_minutes

        if total_minutes > 0:
            excel_time = total_minutes / 1440.0
            cell = ws_edit.cell(row=TOTAL_ROW, column=col)
            cell.value = excel_time
            cell.number_format = 'h:mm;@'

    return totals

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <workbook.xlsm> <input.txt>")
        sys.exit(1)

    wb_path = sys.argv[1]
    input_path = sys.argv[2]

    if not os.path.exists(wb_path):
        print(f"Error: workbook not found: {wb_path}")
        sys.exit(1)
    if not os.path.exists(input_path):
        print(f"Error: input file not found: {input_path}")
        sys.exit(1)

    print(f"Loading input from {input_path}...")
    sessions, comments, notes = parse_input(input_path)
    print(f"  Parsed: {len(sessions)} sessions, {len(comments)} comments, {len(notes)} notes")

    print(f"Loading workbook (data_only) from {wb_path}...")
    wb_data = openpyxl.load_workbook(wb_path, data_only=True)
    ws_data = wb_data[SHEET_NAME]

    print(f"Loading workbook (formulas + VBA) from {wb_path}...")
    wb_edit = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)
    ws_edit = wb_edit[SHEET_NAME]

    # Build maps from the data_only workbook (has computed values)
    print("Building student map...")
    student_map, ambiguous = build_student_map(ws_data)
    print(f"  {len(student_map)} lookup keys, {len(ambiguous)} ambiguous last names: {ambiguous}")

    print("Building date map...")
    date_map = build_date_map(ws_data)
    print(f"  {len(date_map)} dates: {sorted(f'{m}/{d}' for m,d in date_map)}")

    # Fill data into the edit workbook
    print("Filling treatment codes...")
    n_sessions = fill_treatment_codes(ws_edit, sessions, student_map, ambiguous, date_map)
    print(f"  Wrote {n_sessions} cells")

    print("Filling comments...")
    n_comments = fill_comments(ws_edit, comments, student_map, ambiguous)
    print(f"  Wrote {n_comments} comments")

    print("Filling notes...")
    n_dates = fill_notes(ws_edit, notes)
    print(f"  Wrote notes for {n_dates} dates")

    # Get the notes text we just wrote (for daily total calculation)
    notes_text = ws_edit.cell(row=NOTES_ROW, column=1).value or ""
    # Also append A132, A133 if present
    for extra_row in (NOTES_ROW + 1, NOTES_ROW + 2):
        extra = ws_edit.cell(row=extra_row, column=1).value
        if extra:
            notes_text += "\n" + str(extra)

    print("Calculating daily totals...")
    totals = calculate_daily_totals(ws_edit, ws_data, date_map, notes_text)
    for date_str in sorted(totals, key=lambda d: tuple(int(x) for x in d.split("/"))):
        mins = totals[date_str]
        h, m = divmod(int(mins), 60)
        print(f"  {date_str}: {h}:{m:02d} ({mins} min)")

    # Save
    base, ext = os.path.splitext(wb_path)
    out_path = f"{base}_filled{ext}"
    print(f"Saving to {out_path}...")
    wb_edit.save(out_path)
    wb_data.close()
    wb_edit.close()
    print("Done!")


if __name__ == "__main__":
    main()
